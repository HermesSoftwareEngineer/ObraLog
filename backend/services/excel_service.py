"""Excel export for diários de obra using openpyxl."""
from __future__ import annotations

import io
from datetime import date
from typing import Any


_CAMPOS_ORDER = [
    ("estaca_inicial", "Est. Inicial"),
    ("estaca_final", "Est. Final"),
    ("localizacao", "Localização"),
    ("resultado", "Resultado"),
    ("lado_pista", "Lado"),
    ("tempo_manha", "Manhã"),
    ("tempo_tarde", "Tarde"),
]


def _fetch_image_bytes(img: dict, timeout: int = 6) -> bytes | None:
    url = img.get("external_url") or ""
    if url and url.startswith(("http://", "https://")):
        try:
            import urllib.request
            req = urllib.request.Request(url, headers={"User-Agent": "ObraLog/1.0"})
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return resp.read()
        except Exception:
            pass
    storage_path = img.get("storage_path") or ""
    if storage_path:
        try:
            from pathlib import Path
            p = Path(storage_path)
            if p.exists():
                return p.read_bytes()
        except Exception:
            pass
    return None


def _campo_value(r: dict, chave: str) -> Any:
    if chave in ("estaca_inicial", "estaca_final", "resultado"):
        v = r.get(chave)
        return round(float(v), 2) if isinstance(v, (int, float)) else (None if v is None else str(v))
    if chave == "localizacao":
        return str(r.get("localizacao") or "—")
    if chave == "lado_pista":
        return str(r.get("lado_pista") or "—")
    if chave in ("tempo_manha", "tempo_tarde"):
        return str(r.get(chave) or "—")
    return str((r.get("metadata_json") or {}).get(chave) or "—")


def _build_columns(schema: dict) -> list[tuple[str, str]]:
    campos_ativos = schema.get("campos_ativos") or {}
    campos_extras = schema.get("campos_extras") or []
    cols = []
    for chave, label in _CAMPOS_ORDER:
        if campos_ativos.get(chave):
            cols.append((chave, label))
    for extra in campos_extras:
        chave = extra.get("key") or extra.get("chave") or ""
        label = extra.get("label") or chave
        if chave:
            cols.append((chave, label))
    if not cols:
        cols = [("resultado", "Resultado"), ("tempo_manha", "Manhã"), ("tempo_tarde", "Tarde")]
    return cols


def gerar_excel_diario(diario: dict, registros: list[dict], frentes_schemas: dict | None = None) -> bytes:
    """Generate an Excel workbook for a diário de obra. Returns raw xlsx bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    frentes_schemas = frentes_schemas or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    GREEN = "2D6A4F"
    LIGHT_GREEN = "B7E4C7"
    GREY_ROW = "F5F5F5"
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor=GREEN)
    label_font = Font(bold=True, size=10)
    label_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    frente_font = Font(bold=True, size=11, color=GREEN)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    obra_nome = diario.get("obra_nome") or f"Obra {diario.get('obra_id', '')}"
    tipo = diario.get("tipo", "diario").capitalize()
    data_inicio = _fmt_date(str(diario.get("data_inicio", "")))
    data_fim = _fmt_date(str(diario.get("data_fim", "")))
    periodo = f"{data_inicio}" if data_inicio == data_fim else f"{data_inicio} a {data_fim}"
    versao = diario.get("versao_atual", 1)
    status = diario.get("status", "rascunho").capitalize()
    gerado_por = diario.get("gerado_por_nome", "sistema")
    tenant_nome = diario.get("tenant_nome", "")

    # Info header block
    header_data = [
        ("Diário de Obra", f"{tipo} — {obra_nome}"),
        ("Período", periodo),
        ("Versão", str(versao)),
        ("Status", status),
        ("Empresa", tenant_nome),
        ("Gerado por", gerado_por),
    ]
    for row_idx, (label, value) in enumerate(header_data, start=1):
        ws.cell(row=row_idx, column=1, value=label).font = label_font
        ws.cell(row=row_idx, column=1).fill = label_fill
        ws.cell(row=row_idx, column=2, value=value)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 50

    current_row = len(header_data) + 2

    if not registros:
        ws.cell(row=current_row, column=1, value="Nenhum registro no período.")
    else:
        grupos: dict = {}
        for r in registros:
            fid = r.get("frente_servico_id")
            grupos.setdefault(fid, []).append(r)

        for fid, regs in grupos.items():
            schema = frentes_schemas.get(fid, {})
            frente_nome = schema.get("nome") or regs[0].get("frente_servico_nome") or "Sem frente"
            cols = _build_columns(schema)

            col_names = ["#", "Data"] + [label for _, label in cols] + ["Observação", "Status", "Fotos"]
            n_cols = len(col_names)
            foto_col = n_cols  # 1-indexed

            # Frente header row
            frente_cell = ws.cell(row=current_row, column=1, value=f"Frente: {frente_nome}")
            frente_cell.font = frente_font
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=n_cols,
            )
            current_row += 1

            # Column headers
            col_widths = [5, 14] + [max(12, len(label) + 2) for _, label in cols] + [40, 12, 10]
            for ci, (h, w) in enumerate(zip(col_names, col_widths), start=1):
                cell = ws.cell(row=current_row, column=ci, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                ws.column_dimensions[get_column_letter(ci)].width = max(
                    ws.column_dimensions[get_column_letter(ci)].width or 0, w
                )
            current_row += 1

            # Data rows with merged-per-photo pattern
            for seq, r in enumerate(regs, start=1):
                imagens = r.get("imagens") or []
                img_data: list[bytes] = []
                for img in imagens:
                    b = _fetch_image_bytes(img)
                    if b:
                        img_data.append(b)

                n_photo_rows = max(1, len(img_data))
                fill = PatternFill("solid", fgColor=GREY_ROW) if seq % 2 == 0 else PatternFill()

                data_values = (
                    [seq, _fmt_date(str(r.get("data", "")))]
                    + [_campo_value(r, chave) for chave, _ in cols]
                    + [str(r.get("observacao") or ""), str(r.get("status") or "—")]
                )

                row_start = current_row

                for pi in range(n_photo_rows):
                    for ci, val in enumerate(data_values, start=1):
                        cell = ws.cell(row=current_row, column=ci, value=(val if pi == 0 else None))
                        cell.border = border
                        if fill.fgColor and fill.fgColor.value != "00000000":
                            cell.fill = fill

                    # Fotos column: placeholder or image
                    foto_cell = ws.cell(row=current_row, column=foto_col)
                    foto_cell.border = border
                    if fill.fgColor and fill.fgColor.value != "00000000":
                        foto_cell.fill = fill

                    if pi < len(img_data):
                        try:
                            xl_img = XLImage(io.BytesIO(img_data[pi]))
                            xl_img.height = 90
                            xl_img.width = 120
                            ws.row_dimensions[current_row].height = 72
                            ws.add_image(xl_img, get_column_letter(foto_col) + str(current_row))
                        except Exception:
                            foto_cell.value = "[imagem]"
                    elif pi == 0 and not img_data:
                        foto_cell.value = "—"

                    current_row += 1

                # Merge data columns vertically if multiple photo rows
                if n_photo_rows > 1:
                    for ci in range(1, n_cols):  # all except Fotos
                        ws.merge_cells(
                            start_row=row_start, start_column=ci,
                            end_row=current_row - 1, end_column=ci,
                        )
                        merged_cell = ws.cell(row=row_start, column=ci)
                        merged_cell.alignment = Alignment(vertical="center", wrap_text=True)

            current_row += 1  # blank row between frentes

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _fmt_date(value: str) -> str:
    if not value or len(value) < 10:
        return value or ""
    try:
        d = date.fromisoformat(value[:10])
        return d.strftime("%d/%m/%Y")
    except ValueError:
        return value[:10]
